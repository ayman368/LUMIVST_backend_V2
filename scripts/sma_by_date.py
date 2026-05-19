"""Compare close prices BY DATE (not by position) for last 200 trading days."""
import sys, os
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))
from datetime import date as dt_date, datetime, timedelta
from app.core.database import SessionLocal
from app.models.market_pulse import MarketPulse
try:
    import openpyxl
except ImportError:
    import subprocess; subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl

EXCEL_PATH = os.path.join(os.path.dirname(__file__), "..", "..", "market pulse.xlsx")

def to_date(v):
    if isinstance(v, datetime): return v.date()
    if isinstance(v, dt_date): return v
    if isinstance(v, (int, float)):
        return dt_date(1899, 12, 30) + timedelta(days=int(v))
    return None

wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
ws = wb.active
excel_map = {}
for row_idx in range(2, ws.max_row + 1):
    d = to_date(ws.cell(row_idx, 1).value)
    c = ws.cell(row_idx, 5).value
    if d and c: excel_map[d] = float(c)

db = SessionLocal()
rows = db.query(MarketPulse).order_by(MarketPulse.date.desc()).limit(250).all()
db.close()

print(f"DB rows: {len(rows)}, Excel dates: {len(excel_map)}")

# Compare BY DATE
mismatches = 0
for i, r in enumerate(rows):
    db_close = float(r.close)
    ex_close = excel_map.get(r.date)
    if ex_close is None:
        if i < 210:
            print(f"  Row {i}: {r.date} DB={db_close:.2f} -- NOT IN EXCEL")
            mismatches += 1
    elif abs(db_close - ex_close) > 0.02:
        print(f"  Row {i}: {r.date} DB={db_close:.2f} Excel={ex_close:.2f} DIFF={db_close-ex_close:.2f}")
        mismatches += 1

print(f"\nClose price mismatches (by date): {mismatches}")

# Now compute SMA by date-matching
db_dates_sorted = sorted([r.date for r in rows], reverse=True)  # newest first
sma50_db = sum(float(r.close) for r in rows[:50]) / 50
sma150_db = sum(float(r.close) for r in rows[:150]) / 150
sma200_db = sum(float(r.close) for r in rows[:200]) / 200

# Excel SMA using SAME dates as DB
ex_closes_matched = []
for r in rows:
    if r.date in excel_map:
        ex_closes_matched.append(excel_map[r.date])
    else:
        ex_closes_matched.append(float(r.close))  # fallback

sma50_ex = sum(ex_closes_matched[:50]) / 50
sma150_ex = sum(ex_closes_matched[:150]) / 150
sma200_ex = sum(ex_closes_matched[:200]) / 200

print(f"\nSMA (using DB dates, Excel close prices):")
print(f"  SMA50:  {sma50_ex:.2f} vs DB {sma50_db:.2f}  Diff={sma50_db-sma50_ex:.2f}")
print(f"  SMA150: {sma150_ex:.2f} vs DB {sma150_db:.2f}  Diff={sma150_db-sma150_ex:.2f}")
print(f"  SMA200: {sma200_ex:.2f} vs DB {sma200_db:.2f}  Diff={sma200_db-sma200_ex:.2f}")

# Now check: does Excel's positional SMA differ?
# Excel rows sorted newest first
excel_sorted = sorted(excel_map.items(), key=lambda x: x[0], reverse=True)
sma50_ex_pos = sum(c for _, c in excel_sorted[:50]) / 50
sma150_ex_pos = sum(c for _, c in excel_sorted[:150]) / 150
sma200_ex_pos = sum(c for _, c in excel_sorted[:200]) / 200

print(f"\nExcel positional SMA (just sorted by date):")
print(f"  SMA50:  {sma50_ex_pos:.2f}")
print(f"  SMA150: {sma150_ex_pos:.2f}")
print(f"  SMA200: {sma200_ex_pos:.2f}")

# Check if excel has duplicate dates
from collections import Counter
all_dates = []
for row_idx in range(2, ws.max_row + 1):
    d = to_date(ws.cell(row_idx, 1).value)
    if d: all_dates.append(d)
dupes = {d: cnt for d, cnt in Counter(all_dates).items() if cnt > 1}
print(f"\nExcel duplicate dates: {len(dupes)}")
for d, cnt in sorted(dupes.items())[:10]:
    print(f"  {d}: {cnt} times")
