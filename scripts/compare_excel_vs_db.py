"""
Compare DB rows vs Excel rows to find the exact mismatch source.
Reads the Excel file and the DB, then compares row counts and specific close values.
"""
import sys, os
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

from app.core.database import SessionLocal
from app.models.market_pulse import MarketPulse

try:
    import openpyxl
except ImportError:
    print("Installing openpyxl...")
    os.system(f"{sys.executable} -m pip install openpyxl")
    import openpyxl

EXCEL_PATH = os.path.join(os.path.dirname(__file__), "..", "..", "market pulse.xlsx")

def main():
    # --- Read Excel ---
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb.active
    
    excel_rows = []
    for row_idx in range(2, ws.max_row + 1):  # skip header
        date_val = ws.cell(row_idx, 1).value   # A = Date
        close_val = ws.cell(row_idx, 5).value   # E = Close
        ema21_val = ws.cell(row_idx, 12).value  # L = EMA 21
        sma50_val = ws.cell(row_idx, 13).value  # M = SMA 50
        sma150_val = ws.cell(row_idx, 14).value # N = SMA 150
        sma200_val = ws.cell(row_idx, 15).value # O = SMA 200
        if date_val is None:
            break
        excel_rows.append({
            "date": date_val,
            "close": close_val,
            "ema_21": ema21_val,
            "sma_50": sma50_val,
            "sma_150": sma150_val,
            "sma_200": sma200_val,
        })
    
    print(f"Excel total rows: {len(excel_rows)}")
    print(f"Excel newest date: {excel_rows[0]['date']}")
    print(f"Excel oldest date: {excel_rows[-1]['date']}")
    
    # --- Read DB ---
    db = SessionLocal()
    db_records = db.query(MarketPulse).order_by(MarketPulse.date.desc()).all()
    db.close()
    
    print(f"\nDB total rows: {len(db_records)}")
    print(f"DB newest date: {db_records[0].date}")
    print(f"DB oldest date: {db_records[-1].date}")
    
    # --- Build date->close maps ---
    from datetime import date as dt_date, datetime
    
    def to_date(v):
        if isinstance(v, datetime):
            return v.date()
        if isinstance(v, dt_date):
            return v
        return None
    
    excel_map = {}
    for r in excel_rows:
        d = to_date(r["date"])
        if d:
            excel_map[d] = r
    
    db_map = {}
    for r in db_records:
        db_map[r.date] = r
    
    # --- Find dates in Excel but NOT in DB ---
    excel_only = sorted(set(excel_map.keys()) - set(db_map.keys()))
    db_only = sorted(set(db_map.keys()) - set(excel_map.keys()))
    
    print(f"\nDates in Excel but NOT in DB: {len(excel_only)}")
    if excel_only:
        for d in excel_only[:20]:
            print(f"  {d} close={excel_map[d]['close']}")
        if len(excel_only) > 20:
            print(f"  ... and {len(excel_only) - 20} more")
    
    print(f"\nDates in DB but NOT in Excel: {len(db_only)}")
    if db_only:
        for d in db_only[:20]:
            print(f"  {d} close={float(db_map[d].close)}")
        if len(db_only) > 20:
            print(f"  ... and {len(db_only) - 20} more")
    
    # --- Compare latest 5 rows ---
    print("\n--- LATEST 5 COMPARISON ---")
    common_dates = sorted(set(excel_map.keys()) & set(db_map.keys()), reverse=True)
    for d in common_dates[:5]:
        ex = excel_map[d]
        db_r = db_map[d]
        print(f"Date: {d}")
        print(f"  Close:   Excel={ex['close']:.2f}  DB={float(db_r.close):.2f}  Match={abs(float(ex['close']) - float(db_r.close)) < 0.01}")
        if ex['ema_21'] and db_r.ema_21:
            print(f"  EMA21:   Excel={ex['ema_21']:.2f}  DB={float(db_r.ema_21):.2f}  Diff={float(db_r.ema_21) - float(ex['ema_21']):.2f}")
        if ex['sma_50'] and db_r.sma_50:
            print(f"  SMA50:   Excel={ex['sma_50']:.2f}  DB={float(db_r.sma_50):.2f}  Diff={float(db_r.sma_50) - float(ex['sma_50']):.2f}")
        if ex['sma_150'] and db_r.sma_150:
            print(f"  SMA150:  Excel={ex['sma_150']:.2f}  DB={float(db_r.sma_150):.2f}  Diff={float(db_r.sma_150) - float(ex['sma_150']):.2f}")
        if ex['sma_200'] and db_r.sma_200:
            print(f"  SMA200:  Excel={ex['sma_200']:.2f}  DB={float(db_r.sma_200):.2f}  Diff={float(db_r.sma_200) - float(ex['sma_200']):.2f}")
    
    # --- Check close mismatches (shared dates with different close) ---
    mismatched_closes = []
    for d in common_dates:
        ex_close = float(excel_map[d]['close'])
        db_close = float(db_map[d].close)
        if abs(ex_close - db_close) > 0.01:
            mismatched_closes.append((d, ex_close, db_close))
    
    print(f"\n--- CLOSE PRICE MISMATCHES (shared dates): {len(mismatched_closes)} ---")
    for d, ex_c, db_c in mismatched_closes[:10]:
        print(f"  {d}: Excel={ex_c:.2f}  DB={db_c:.2f}  Diff={db_c - ex_c:.2f}")

if __name__ == "__main__":
    main()
