"""
Pinpoint which close prices differ between DB and Excel in the SMA window.
Compare the last 200 close prices row-by-row.
"""
import sys, os
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

from datetime import date as dt_date, datetime, timedelta
from app.core.database import SessionLocal
from app.models.market_pulse import MarketPulse

try:
    import openpyxl
except ImportError:
    os.system(f"{sys.executable} -m pip install openpyxl")
    import openpyxl

EXCEL_PATH = os.path.join(os.path.dirname(__file__), "..", "..", "market pulse.xlsx")


def to_date(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, dt_date):
        return v
    if isinstance(v, (int, float)):
        epoch = dt_date(1899, 12, 30)
        return epoch + timedelta(days=int(v))
    return None


def main():
    # --- Excel: get rows newest-first (row 2 = newest) ---
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb.active
    
    excel_list = []
    for row_idx in range(2, ws.max_row + 1):
        date_val = ws.cell(row_idx, 1).value
        close_val = ws.cell(row_idx, 5).value
        if date_val is None:
            break
        d = to_date(date_val)
        if d and close_val is not None:
            excel_list.append((d, float(close_val)))
    
    # Excel is already newest-first (row 2 = most recent)
    print(f"Excel: {len(excel_list)} rows, newest={excel_list[0][0]}, oldest={excel_list[-1][0]}")
    
    # --- DB: get rows newest-first ---
    db = SessionLocal()
    db_records = db.query(MarketPulse).order_by(MarketPulse.date.desc()).limit(250).all()
    db.close()
    
    db_list = [(r.date, float(r.close)) for r in db_records]
    print(f"DB: {len(db_list)} rows, newest={db_list[0][0]}, oldest={db_list[-1][0]}")
    
    # --- Row-by-row comparison (positional, like Excel SMA) ---
    print("\n--- ROW-BY-ROW COMPARISON (newest first) ---")
    print(f"{'Row':>4} | {'Excel Date':>12} {'Excel Close':>12} | {'DB Date':>12} {'DB Close':>12} | {'Match':>5}")
    print("-" * 80)
    
    mismatches = 0
    for i in range(min(210, len(excel_list), len(db_list))):
        ex_d, ex_c = excel_list[i]
        db_d, db_c = db_list[i]
        
        date_match = ex_d == db_d
        close_match = abs(ex_c - db_c) < 0.02
        
        if not date_match or not close_match:
            marker = "<<< DIFF"
            mismatches += 1
        else:
            marker = ""
        
        # Print all rows 0-5, then only mismatches, plus rows around SMA boundaries
        if i < 5 or not date_match or not close_match or i in (49, 50, 51, 149, 150, 151, 199, 200):
            print(f"{i:>4} | {str(ex_d):>12} {ex_c:>12.2f} | {str(db_d):>12} {db_c:>12.2f} | {marker}")
    
    print(f"\nTotal mismatches in first 210 rows: {mismatches}")
    
    # --- Compute SMA manually from both sources ---
    if len(excel_list) >= 200 and len(db_list) >= 200:
        ex_sma50 = sum(c for _, c in excel_list[:50]) / 50
        db_sma50 = sum(c for _, c in db_list[:50]) / 50
        ex_sma150 = sum(c for _, c in excel_list[:150]) / 150
        db_sma150 = sum(c for _, c in db_list[:150]) / 150
        ex_sma200 = sum(c for _, c in excel_list[:200]) / 200
        db_sma200 = sum(c for _, c in db_list[:200]) / 200
        
        print(f"\n--- MANUAL SMA CHECK ---")
        print(f"SMA 50:  Excel={ex_sma50:.2f}  DB={db_sma50:.2f}  Diff={db_sma50-ex_sma50:.2f}")
        print(f"SMA 150: Excel={ex_sma150:.2f}  DB={db_sma150:.2f}  Diff={db_sma150-ex_sma150:.2f}")
        print(f"SMA 200: Excel={ex_sma200:.2f}  DB={db_sma200:.2f}  Diff={db_sma200-ex_sma200:.2f}")


if __name__ == "__main__":
    main()
