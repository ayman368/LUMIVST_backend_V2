"""Find extra dates in historical_reports (2025+) that are NOT in Excel."""
import sys, os
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))
from datetime import date as dt_date, datetime, timedelta
from app.core.database import SessionLocal
from app.models.market_reports import HistoricalReport
try:
    import openpyxl
except ImportError:
    os.system(f"{sys.executable} -m pip install openpyxl")
    import openpyxl

EXCEL_PATH = os.path.join(os.path.dirname(__file__), "..", "..", "market pulse.xlsx")

def to_date(v):
    if isinstance(v, datetime): return v.date()
    if isinstance(v, dt_date): return v
    if isinstance(v, (int, float)):
        return dt_date(1899, 12, 30) + timedelta(days=int(v))
    return None

wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
ws = wb.active
excel_dates = set()
for row_idx in range(2, ws.max_row + 1):
    d = to_date(ws.cell(row_idx, 1).value)
    if d: excel_dates.add(d)

db = SessionLocal()
# Check 2025+ specifically
rows = db.query(HistoricalReport).filter(
    HistoricalReport.report_date >= dt_date(2025, 1, 1)
).order_by(HistoricalReport.report_date.asc()).all()
db.close()

db_dates_2025 = set(r.report_date for r in rows)
excel_dates_2025 = {d for d in excel_dates if d >= dt_date(2025, 1, 1)}

extra_in_db = sorted(db_dates_2025 - excel_dates_2025)
missing_in_db = sorted(excel_dates_2025 - db_dates_2025)

print(f"DB rows 2025+: {len(db_dates_2025)}")
print(f"Excel rows 2025+: {len(excel_dates_2025)}")
print(f"\nExtra in DB (not in Excel): {len(extra_in_db)}")
for d in extra_in_db[:30]:
    print(f"  {d} ({d.strftime('%A')})")
print(f"\nMissing in DB (in Excel only): {len(missing_in_db)}")
for d in missing_in_db[:10]:
    print(f"  {d}")
